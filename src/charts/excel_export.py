"""수식이 살아있는 DCF Excel 모델 생성 (IB 실무 레이아웃)"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1

# ── 색상 상수 ──────────────────────────────────────────────────────────────
NAVY   = "003366"
LBLUE  = "D6E4F0"
YELLOW = "FFF2CC"
GREEN  = "E2EFDA"
GRAY   = "F2F2F2"
WHITE  = "FFFFFF"
RED    = "C00000"


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thin = Side(style="thin",   color="CCCCCC")
    med  = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=med)

def _pct_fmt():  return "0.0%"
def _num_fmt():  return '#,##0'
def _num1_fmt(): return '#,##0.0'

def _header_row(ws, row, cols_text, bg=NAVY, fg="FFFFFF", bold=True):
    for col, text in enumerate(cols_text, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font    = _font(bold=bold, color=fg, size=10)
        c.fill    = _fill(bg)
        c.border  = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _label(ws, row, col, text, bold=False, bg=GRAY):
    c = ws.cell(row=row, column=col, value=text)
    c.font   = _font(bold=bold)
    c.fill   = _fill(bg)
    c.border = _border()
    c.alignment = Alignment(horizontal="left", vertical="center")

def _val(ws, row, col, value, fmt=_num_fmt(), bold=False, bg=WHITE, formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font         = _font(bold=bold)
    c.fill         = _fill(bg)
    c.border       = _border()
    c.number_format = fmt
    c.alignment    = Alignment(horizontal="right", vertical="center")
    return c

def _input(ws, row, col, value, fmt=_num_fmt()):
    """편집 가능한 입력 셀 (노란색)"""
    c = _val(ws, row, col, value, fmt=fmt, bg=YELLOW)
    return c


def build_excel_model(
    corp_name: str,
    fin_df: pd.DataFrame,
    assumptions,          # DCFAssumptions
    dcf_result: dict,
    net_debt: float,
    mkt_cap: float,
    segments: list = None,  # [{"name": str, "revenue": float, "growths": [g1..g5]}]
) -> io.BytesIO:
    """
    완전한 수식 DCF Excel 모델 반환
    sheets: Assumptions | Segments(optional) | DCF Model | Sensitivity | Summary
    """
    wb = Workbook()

    # ── Sheet 1: Assumptions ──────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "Assumptions"
    _build_assumptions(ws_a, corp_name, assumptions, net_debt, mkt_cap)

    # ── Sheet 2: Segments (사업부별 매출 추정) ──────────────────────────
    if segments:
        ws_s = wb.create_sheet("Segments")
        _build_segments(ws_s, segments, assumptions)

    # ── Sheet 3: DCF Model ───────────────────────────────────────────────
    ws_d = wb.create_sheet("DCF Model")
    seg_sheet = "Segments" if segments else None
    _build_dcf(ws_d, fin_df, assumptions, net_debt, seg_sheet)

    # ── Sheet 4: Sensitivity ─────────────────────────────────────────────
    ws_sens = wb.create_sheet("Sensitivity")
    _build_sensitivity(ws_sens, dcf_result, assumptions)

    # ── Sheet 5: Summary ─────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _build_summary(ws_sum, corp_name, dcf_result, mkt_cap, assumptions)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
def _build_assumptions(ws, corp_name, a, net_debt, mkt_cap):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20

    # 타이틀
    t = ws.cell(row=1, column=1, value=f"{corp_name} — DCF 가정 (Assumptions)")
    t.font = _font(bold=True, color=NAVY, size=13)
    ws.merge_cells("A1:C1")

    ws.cell(row=2, column=1, value="※ 노란색 셀은 편집 가능합니다. DCF Model 시트의 수식이 이 셀을 참조합니다.")
    ws.cell(row=2, column=1).font = _font(color="888888", size=9)

    rows = [
        (4,  "── WACC 구성 ──",              None,   None,   NAVY, "FFFFFF"),
        (5,  "WACC",                          a.wacc, _pct_fmt(), None, None),
        (6,  "Risk-Free Rate (Rf)",           0.035,  _pct_fmt(), None, None),
        (7,  "Equity Risk Premium (ERP)",     0.06,   _pct_fmt(), None, None),
        (8,  "Beta",                          1.0,    _num1_fmt(), None, None),
        (9,  "Cost of Debt (Kd)",             0.05,   _pct_fmt(), None, None),
        (10, "Tax Rate",                      a.tax_rate, _pct_fmt(), None, None),
        (12, "── Terminal Value ──",           None,   None,   NAVY, "FFFFFF"),
        (13, "Terminal Growth Rate (TGR)",    a.terminal_growth_rate, _pct_fmt(), None, None),
        (14, "Terminal Method",               "Gordon Growth" if a.terminal_method=="gordon" else "Exit Multiple", "@", None, None),
        (15, "Exit EV/EBITDA Multiple",       a.exit_multiple, _num1_fmt(), None, None),
        (17, "── 매출 성장률 가정 ──",          None,   None,   NAVY, "FFFFFF"),
        (18, "Year 1 Revenue Growth",         a.revenue_growth_rates[0], _pct_fmt(), None, None),
        (19, "Year 2 Revenue Growth",         a.revenue_growth_rates[1] if len(a.revenue_growth_rates)>1 else 0, _pct_fmt(), None, None),
        (20, "Year 3 Revenue Growth",         a.revenue_growth_rates[2] if len(a.revenue_growth_rates)>2 else 0, _pct_fmt(), None, None),
        (21, "Year 4 Revenue Growth",         a.revenue_growth_rates[3] if len(a.revenue_growth_rates)>3 else 0, _pct_fmt(), None, None),
        (22, "Year 5 Revenue Growth",         a.revenue_growth_rates[4] if len(a.revenue_growth_rates)>4 else 0, _pct_fmt(), None, None),
        (24, "── P&L 가정 ──",                None,   None,   NAVY, "FFFFFF"),
        (25, "EBITDA Margin",                 a.ebitda_margin, _pct_fmt(), None, None),
        (26, "D&A / Revenue",                 a.da_pct_revenue, _pct_fmt(), None, None),
        (27, "CapEx / Revenue",               a.capex_pct_revenue, _pct_fmt(), None, None),
        (28, "NWC Change / ΔRevenue",         a.nwc_change_pct, _pct_fmt(), None, None),
        (30, "── Balance Sheet ──",           None,   None,   NAVY, "FFFFFF"),
        (31, "Net Debt (억원)",               net_debt, _num_fmt(), None, None),
        (32, "현재 시가총액 (억원)",           mkt_cap,  _num_fmt(), None, None),
    ]

    for row, label, value, fmt, bg_override, fg_override in rows:
        if bg_override:
            _header_row(ws, row, [label, "", ""], bg=bg_override, fg=fg_override)
        else:
            _label(ws, row, 1, label, bold=False)
            if value is not None:
                _input(ws, row, 2, value, fmt=fmt or _num_fmt())
                _label(ws, row, 3, "← 편집 가능", bold=False, bg=YELLOW)


# ─────────────────────────────────────────────────────────────────────────────
def _build_segments(ws, segments, a):
    ws.column_dimensions["A"].width = 22
    for i in range(8):
        ws.column_dimensions[get_column_letter(i+2)].width = 14

    _header_row(ws, 1, ["사업부 (Segment)", "Base 매출(억)", "Y1 성장률", "Y2 성장률",
                         "Y3 성장률", "Y4 성장률", "Y5 성장률",
                         "Y1 매출", "Y2 매출", "Y3 매출", "Y4 매출", "Y5 매출"])

    n = len(segments)
    for i, seg in enumerate(segments):
        r = i + 2
        growths = seg.get("growths", [0]*5)
        while len(growths) < 5:
            growths.append(0)

        _input(ws, r, 1, seg["name"], fmt="@")
        _input(ws, r, 2, seg["revenue"])
        for j, g in enumerate(growths[:5]):
            _input(ws, r, 3+j, g, fmt=_pct_fmt())

        # 수식: Y1 매출 = Base × (1+g1), Y2 = Y1 × (1+g2) ...
        base_col = get_column_letter(2)
        for j in range(5):
            g_col  = get_column_letter(3 + j)
            if j == 0:
                formula = f"=B{r}*(1+C{r})"
            else:
                prev   = get_column_letter(8 + j - 1)
                g_col2 = get_column_letter(3 + j)
                formula = f"={prev}{r}*(1+{g_col2}{r})"
            _val(ws, r, 8+j, formula, fmt=_num_fmt(), bg=GREEN)

    # 합계 행
    total_row = n + 2
    _label(ws, total_row, 1, "합계 (Total Revenue)", bold=True, bg=LBLUE)
    _val(ws, total_row, 2,
         f"=SUM(B2:B{n+1})", fmt=_num_fmt(), bold=True, bg=LBLUE)
    for j in range(5):
        c = get_column_letter(8+j)
        _val(ws, total_row, 8+j,
             f"=SUM({c}2:{c}{n+1})", fmt=_num_fmt(), bold=True, bg=LBLUE)


# ─────────────────────────────────────────────────────────────────────────────
def _build_dcf(ws, fin_df, a, net_debt, seg_sheet=None):
    proj = a.projection_years
    years = ["Base"] + [f"Y+{i+1}" for i in range(proj)]
    col_offset = 2  # A=label, B=Base, C=Y1 ...

    ws.column_dimensions["A"].width = 30
    for i in range(proj + 2):
        ws.column_dimensions[get_column_letter(i+2)].width = 14

    # 타이틀
    t = ws.cell(row=1, column=1, value="DCF Model — Free Cash Flow Projection")
    t.font = _font(bold=True, color=NAVY, size=12)

    # 컬럼 헤더
    _header_row(ws, 3, ["항목 (억원)"] + years)

    # 기준연도 실적
    latest = fin_df.iloc[-1] if not fin_df.empty else {}
    def g(key): return float(latest.get(key) or 0) if hasattr(latest, 'get') else 0

    base_rev    = g("revenue")
    base_ebitda = g("operating_income") + (g("depreciation") or base_rev * 0.05)
    base_da     = g("depreciation") or base_rev * 0.05

    ASMP = "Assumptions"  # 참조 시트명

    # ── 매출액 행 ──────────────────────────────────────────────────────────
    ROW_REV = 5
    _label(ws, ROW_REV, 1, "매출액 (Revenue)", bold=True, bg=LBLUE)
    _val(ws, ROW_REV, 2, base_rev, fmt=_num_fmt(), bg=LBLUE, bold=True)
    for i in range(1, proj+1):
        col     = get_column_letter(i + col_offset)
        prev    = get_column_letter(i + col_offset - 1)
        g_row   = 17 + i  # Assumptions 시트의 성장률 행
        if seg_sheet:
            # 세그먼트 합계 참조
            seg_col = get_column_letter(7 + i)
            ws.cell(row=ROW_REV, column=i+col_offset,
                    value=f"='{seg_sheet}'!{seg_col}{len([1])+2}")
            ws.cell(row=ROW_REV, column=i+col_offset).number_format = _num_fmt()
            ws.cell(row=ROW_REV, column=i+col_offset).fill   = _fill(GREEN)
            ws.cell(row=ROW_REV, column=i+col_offset).border = _border()
        else:
            f = f"={prev}{ROW_REV}*(1+'{ASMP}'!$B${g_row})"
            _val(ws, ROW_REV, i+col_offset, f, fmt=_num_fmt(), bg=GREEN)

    # ── 성장률 행 ──────────────────────────────────────────────────────────
    ROW_GR = 6
    _label(ws, ROW_GR, 1, "  Revenue Growth (%)", bg=GRAY)
    _val(ws, ROW_GR, 2, "-", fmt="@", bg=GRAY)
    for i in range(1, proj+1):
        col   = get_column_letter(i + col_offset)
        prev  = get_column_letter(i + col_offset - 1)
        f = f"=({col}{ROW_REV}-{prev}{ROW_REV})/{prev}{ROW_REV}"
        _val(ws, ROW_GR, i+col_offset, f, fmt=_pct_fmt(), bg=GRAY)

    # ── EBITDA ────────────────────────────────────────────────────────────
    ROW_EBITDA = 8
    _label(ws, ROW_EBITDA, 1, "EBITDA", bold=True, bg=LBLUE)
    _val(ws, ROW_EBITDA, 2, base_ebitda, fmt=_num_fmt(), bg=LBLUE, bold=True)
    for i in range(1, proj+1):
        col = get_column_letter(i + col_offset)
        f   = f"={col}{ROW_REV}*'{ASMP}'!$B$25"
        _val(ws, ROW_EBITDA, i+col_offset, f, fmt=_num_fmt(), bg=LBLUE)

    # ── D&A ───────────────────────────────────────────────────────────────
    ROW_DA = 9
    _label(ws, ROW_DA, 1, "  (-)  D&A", bg=GRAY)
    _val(ws, ROW_DA, 2, base_da, fmt=_num_fmt(), bg=GRAY)
    for i in range(1, proj+1):
        col = get_column_letter(i + col_offset)
        f   = f"={col}{ROW_REV}*'{ASMP}'!$B$26"
        _val(ws, ROW_DA, i+col_offset, f, fmt=_num_fmt(), bg=GRAY)

    # ── EBIT ──────────────────────────────────────────────────────────────
    ROW_EBIT = 10
    _label(ws, ROW_EBIT, 1, "EBIT", bold=True)
    _val(ws, ROW_EBIT, 2, base_ebitda - base_da, fmt=_num_fmt(), bold=True)
    for i in range(1, proj+1):
        col = get_column_letter(i + col_offset)
        f   = f"={col}{ROW_EBITDA}-{col}{ROW_DA}"
        _val(ws, ROW_EBIT, i+col_offset, f, fmt=_num_fmt(), bold=True)

    # ── NOPAT ─────────────────────────────────────────────────────────────
    ROW_NOPAT = 11
    _label(ws, ROW_NOPAT, 1, "  NOPAT = EBIT × (1-t)", bg=GRAY)
    for i in range(col_offset, proj+col_offset+1):
        col = get_column_letter(i)
        f   = f"={col}{ROW_EBIT}*(1-'{ASMP}'!$B$10)"
        _val(ws, ROW_NOPAT, i, f, fmt=_num_fmt(), bg=GRAY)

    # ── CapEx ─────────────────────────────────────────────────────────────
    ROW_CAPEX = 12
    _label(ws, ROW_CAPEX, 1, "  (-) CapEx", bg=GRAY)
    for i in range(col_offset, proj+col_offset+1):
        col = get_column_letter(i)
        f   = f"={col}{ROW_REV}*'{ASMP}'!$B$27"
        _val(ws, ROW_CAPEX, i, f, fmt=_num_fmt(), bg=GRAY)

    # ── ΔNWC ─────────────────────────────────────────────────────────────
    ROW_NWC = 13
    _label(ws, ROW_NWC, 1, "  (-) ΔNWC", bg=GRAY)
    _val(ws, ROW_NWC, 2, 0, fmt=_num_fmt(), bg=GRAY)
    for i in range(1, proj+1):
        col  = get_column_letter(i + col_offset)
        prev = get_column_letter(i + col_offset - 1)
        f    = f"=({col}{ROW_REV}-{prev}{ROW_REV})*'{ASMP}'!$B$28"
        _val(ws, ROW_NWC, i+col_offset, f, fmt=_num_fmt(), bg=GRAY)

    # ── Free Cash Flow ────────────────────────────────────────────────────
    ROW_FCF = 15
    _label(ws, ROW_FCF, 1, "Free Cash Flow (FCF)", bold=True, bg=NAVY)
    ws.cell(row=ROW_FCF, column=1).font = _font(bold=True, color="FFFFFF")
    for i in range(col_offset, proj+col_offset+1):
        col = get_column_letter(i)
        f   = f"={col}{ROW_NOPAT}+{col}{ROW_DA}-{col}{ROW_CAPEX}-{col}{ROW_NWC}"
        _val(ws, ROW_FCF, i, f, fmt=_num_fmt(), bold=True, bg=LBLUE)

    # ── Discount Factor ────────────────────────────────────────────────────
    ROW_DF = 16
    _label(ws, ROW_DF, 1, "  Discount Factor  1/(1+WACC)^t", bg=GRAY)
    _val(ws, ROW_DF, 2, 1.0, fmt="0.0000", bg=GRAY)
    for i in range(1, proj+1):
        col = get_column_letter(i + col_offset)
        f   = f"=1/(1+'{ASMP}'!$B$5)^{i}"
        _val(ws, ROW_DF, i+col_offset, f, fmt="0.0000", bg=GRAY)

    # ── PV(FCF) ────────────────────────────────────────────────────────────
    ROW_PVFCF = 17
    _label(ws, ROW_PVFCF, 1, "  PV (FCF)", bg=GRAY)
    _val(ws, ROW_PVFCF, 2, "-", fmt="@", bg=GRAY)
    for i in range(1, proj+1):
        col = get_column_letter(i + col_offset)
        f   = f"={col}{ROW_FCF}*{col}{ROW_DF}"
        _val(ws, ROW_PVFCF, i+col_offset, f, fmt=_num_fmt(), bg=GRAY)

    # ── Terminal Value & Bridge ────────────────────────────────────────────
    last_col = get_column_letter(proj + col_offset)

    ROW_SUM    = 19
    ROW_TV     = 20
    ROW_PVTV   = 21
    ROW_EV     = 23
    ROW_NETD   = 24
    ROW_EQ     = 25

    pv_cols = "".join(
        f"{get_column_letter(i+col_offset)}{ROW_PVFCF}+"
        for i in range(1, proj+1)
    ).rstrip("+")

    _label(ws, ROW_SUM,  1, "∑ PV(FCF)",              bold=True, bg=LBLUE)
    _label(ws, ROW_TV,   1, "Terminal Value (Gordon)", bold=True, bg=LBLUE)
    _label(ws, ROW_PVTV, 1, "PV(Terminal Value)",      bold=True, bg=LBLUE)
    _label(ws, ROW_EV,   1, "Enterprise Value (EV)",   bold=True, bg=NAVY)
    ws.cell(row=ROW_EV, column=1).font = _font(bold=True, color="FFFFFF")
    _label(ws, ROW_NETD, 1, "(-) Net Debt (순차입금)")
    _label(ws, ROW_EQ,   1, "Equity Value (주주가치)", bold=True, bg=NAVY)
    ws.cell(row=ROW_EQ, column=1).font = _font(bold=True, color="FFFFFF")

    _val(ws, ROW_SUM,  2, f"={pv_cols}", fmt=_num_fmt(), bold=True, bg=LBLUE)
    _val(ws, ROW_TV,   2,
         f"={last_col}{ROW_FCF}*(1+'{ASMP}'!$B$13)/('{ASMP}'!$B$5-'{ASMP}'!$B$13)",
         fmt=_num_fmt(), bold=True, bg=LBLUE)
    _val(ws, ROW_PVTV, 2,
         f"=B{ROW_TV}/(1+'{ASMP}'!$B$5)^{proj}",
         fmt=_num_fmt(), bold=True, bg=LBLUE)
    _val(ws, ROW_EV,   2, f"=B{ROW_SUM}+B{ROW_PVTV}", fmt=_num_fmt(), bold=True, bg=LBLUE)
    _val(ws, ROW_NETD, 2, net_debt,            fmt=_num_fmt())
    _val(ws, ROW_EQ,   2, f"=B{ROW_EV}-B{ROW_NETD}", fmt=_num_fmt(), bold=True, bg=GREEN)

    # TV 비중
    _label(ws, ROW_PVTV + 1, 1, "  TV / EV 비중")
    _val(ws, ROW_PVTV+1, 2, f"=B{ROW_PVTV}/B{ROW_EV}", fmt=_pct_fmt(), bg=GRAY)


# ─────────────────────────────────────────────────────────────────────────────
def _build_sensitivity(ws, dcf_result, a):
    ws.column_dimensions["A"].width = 16
    for i in range(7):
        ws.column_dimensions[get_column_letter(i+2)].width = 14

    t = ws.cell(row=1, column=1, value="민감도 분석 — Equity Value (억원)")
    t.font = _font(bold=True, color=NAVY, size=12)

    ws.cell(row=2, column=1, value="행: Terminal Growth Rate  /  열: WACC").font = _font(color="666666", size=9)

    waccs = [a.wacc - 0.02, a.wacc - 0.01, a.wacc, a.wacc + 0.01, a.wacc + 0.02]
    tgrs  = [a.terminal_growth_rate - 0.01, a.terminal_growth_rate,
             a.terminal_growth_rate + 0.005, a.terminal_growth_rate + 0.01,
             a.terminal_growth_rate + 0.015]

    # 헤더
    _label(ws, 4, 1, "TGR \\ WACC", bold=True, bg=NAVY)
    ws.cell(row=4, column=1).font = _font(bold=True, color="FFFFFF")
    for j, w in enumerate(waccs):
        c = ws.cell(row=4, column=j+2, value=w)
        c.number_format = _pct_fmt()
        c.font   = _font(bold=True, color="FFFFFF")
        c.fill   = _fill(NAVY)
        c.border = _border()
        c.alignment = Alignment(horizontal="center")

    base_ev = dcf_result["equity_value"]

    from src.models.dcf import DCFAssumptions, run_dcf
    base_rev    = dcf_result.get("_base_rev", 0)
    base_ebitda = dcf_result.get("_base_ebitda", 0)
    net_debt    = dcf_result.get("_net_debt", 0)

    for i, tgr in enumerate(tgrs):
        r = i + 5
        c = ws.cell(row=r, column=1, value=tgr)
        c.number_format = _pct_fmt()
        c.font   = _font(bold=True)
        c.fill   = _fill(GRAY)
        c.border = _border()

        for j, wacc in enumerate(waccs):
            try:
                tmp = DCFAssumptions(**{**a.__dict__, "wacc": wacc, "terminal_growth_rate": tgr})
                val = run_dcf(base_rev, base_ebitda, net_debt, tmp)["equity_value"]
            except Exception:
                val = None

            is_base = (abs(wacc - a.wacc) < 1e-6 and abs(tgr - a.terminal_growth_rate) < 1e-6)
            bg = NAVY if is_base else (GREEN if val and val > base_ev * 1.1 else
                                       ("FFC7CE" if val and val < base_ev * 0.9 else WHITE))
            fg = "FFFFFF" if is_base else "000000"
            cell = ws.cell(row=r, column=j+2, value=val)
            cell.number_format = _num_fmt()
            cell.fill   = _fill(bg)
            cell.border = _border()
            cell.font   = _font(bold=is_base, color=fg)
            cell.alignment = Alignment(horizontal="right")


# ─────────────────────────────────────────────────────────────────────────────
def _build_summary(ws, corp_name, dcf_result, mkt_cap, a):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20

    t = ws.cell(row=1, column=1, value=f"{corp_name} — Valuation Summary")
    t.font = _font(bold=True, color=NAVY, size=14)
    ws.merge_cells("A1:C1")

    ev  = dcf_result.get("equity_value", 0) + dcf_result.get("_net_debt", 0)
    eq  = dcf_result.get("equity_value", 0)
    upside = (eq - mkt_cap) / mkt_cap * 100 if mkt_cap > 0 else None

    rows = [
        (3,  "── DCF 결과 ──",               None,    None,    NAVY, "FFFFFF"),
        (4,  "Enterprise Value (EV)",         ev,      _num_fmt(), None, None),
        (5,  "Equity Value",                  eq,      _num_fmt(), None, None),
        (6,  "현재 시가총액",                  mkt_cap, _num_fmt(), None, None),
        (7,  "Upside / Downside",             upside,  "0.0%",    None, None),
        (8,  "WACC",                          a.wacc,  _pct_fmt(), None, None),
        (9,  "Terminal Growth Rate",          a.terminal_growth_rate, _pct_fmt(), None, None),
        (10, "TV / EV 비중",                  dcf_result.get("tv_pct", 0)/100, _pct_fmt(), None, None),
    ]

    for row, label, value, fmt, bg_override, fg_override in rows:
        if bg_override:
            _header_row(ws, row, [label, "", ""], bg=bg_override, fg=fg_override)
        else:
            _label(ws, row, 1, label)
            if value is not None:
                c = ws.cell(row=row, column=2, value=value)
                c.number_format = fmt or _num_fmt()
                c.border = _border()
                c.alignment = Alignment(horizontal="right")
                # upside 색상
                if label.startswith("Upside") and upside is not None:
                    c.font = _font(bold=True,
                                   color="00A020" if upside >= 10 else
                                   ("CC0000" if upside <= -10 else "BB8800"))
